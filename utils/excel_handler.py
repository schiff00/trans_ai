import pandas as pd
from openpyxl import load_workbook
import os
import shutil

class ExcelHandler:
    def __init__(self, logger):
        self.logger = logger

    def get_sheet_names(self, file_path):
        return pd.ExcelFile(file_path).sheet_names

    def load_sheet(self, file_path, sheet_name):
        return pd.read_excel(file_path, sheet_name=sheet_name)

    def save_as_text_file(self, excel_file, df, target_column):
        try:
            base_name, _ = os.path.splitext(excel_file)
            new_file_name = f"{base_name}_trans.txt"

            self.logger.info(f"새 텍스트 파일로 저장합니다: {new_file_name}")

            with open(new_file_name, 'w', encoding='utf-8') as file:
                for index, row in df.iterrows():
                    korean_text = row.get("m_Talk_KOREA", "")
                    speaker = row.get("m_CharStr", "Unknown Speaker")
                    translated_text = row.get(target_column, "")

                    if pd.notna(korean_text) and pd.notna(translated_text):
                        file.write(f"{speaker}:\n")
                        file.write(f"  한국어: {korean_text}\n")
                        file.write(f"  번역: {translated_text}\n\n")

            self.logger.info(f"텍스트 파일이 성공적으로 저장되었습니다: {new_file_name}")
            return new_file_name
        except Exception as e:
            self.logger.error(f"텍스트 파일을 저장하는데 오류가 발생했습니다: {str(e)}")
            raise

    def save_as_excel_file(self, excel_file, df, current_sheet, target_column, translated_items):
        try:
            base_name, ext = os.path.splitext(excel_file)
            new_file_name = f"{base_name}_translated{ext}"
            shutil.copy2(excel_file, new_file_name)

            self.logger.info(f"원본 파일을 복사했습니다: {new_file_name}")

            workbook = load_workbook(new_file_name)
            sheet = workbook[current_sheet]

            header_row = None
            for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                if "m_CutScenKey" in row:
                    header_row = row
                    break

            if header_row is None:
                raise ValueError("헤더 행을 찾을 수 없습니다.")

            cut_scen_key_index = header_row.index("m_CutScenKey")
            lang_column_index = header_row.index(target_column) if target_column in header_row else None

            if lang_column_index is None:
                lang_column_index = len(header_row)
                sheet.cell(row=1, column=lang_column_index + 1, value=target_column)

            for item in translated_items:
                row = df.loc[df['m_CutScenKey'] == item].iloc[0]
                cut_scen_key = row["m_CutScenKey"]
                translated_text = row[target_column]

                for excel_row in sheet.iter_rows(min_row=2, values_only=False):
                    if excel_row[cut_scen_key_index].value == cut_scen_key:
                        excel_row[lang_column_index].value = translated_text
                        break

            workbook.save(new_file_name)

            self.logger.info(f"번역된 내용이 엑셀 파일에 저장되었습니다: {new_file_name}")
            return new_file_name
        except Exception as e:
            self.logger.error(f"엑셀 파일을 저장하는데 오류가 발생했습니다: {str(e)}")
            raise